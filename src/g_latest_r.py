import os
import sqlite3
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.chart import PieChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, Reference

# Load environment variables
load_dotenv()

# Get Telegram bot token and allowed user ID from .env
API_TOKEN = os.getenv("API_TOKEN")
ALLOWED_USER_ID = os.getenv("ALLOWED_USER_ID")
DB_PATH = os.getenv("DB_PATH")

# Function to fetch data from SQLite database
def fetch_data_from_db(db_path):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    # Fetching all data from the transactions table
    cursor.execute("SELECT * FROM transactions")
    rows = cursor.fetchall()
    conn.close()
    return rows

# Function to auto-resize columns based on content length
def auto_resize_columns(ws, data):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Function to add borders to all cells in the worksheet
def add_borders(ws):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

def generate_excel_report(data, report_file):
    wb = Workbook()
    headers = ["id", "type", "category", "quantity", "amount", "description", "created_at", "is_outlier"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Month", "Income", "Expense (Outlier Excluded)", "Expense Outlier"])

    # Border helper
    def add_borders(ws):
        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin

    # Auto-resize helper
    def auto_resize_columns(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    # Data aggregation
    month_data = {}
    monthly_income_expense = {}

    for row in data:
        created_at = datetime.strptime(row[6], "%Y-%m-%d %H:%M:%S")
        month_str = created_at.strftime("%Y%m")

        if month_str not in month_data:
            month_data[month_str] = []
            monthly_income_expense[month_str] = {
                "income": 0,
                "expense_clean": 0,
                "expense_outlier": 0
            }

        month_data[month_str].append(row)

        amount = row[4]
        is_outlier = row[7] if row[7] is not None else 0

        if row[1] == 'expense':
            if is_outlier == 1:
                monthly_income_expense[month_str]["expense_outlier"] += amount
            else:
                monthly_income_expense[month_str]["expense_clean"] += amount
        else:
            monthly_income_expense[month_str]["income"] += amount

    # Write summary data
    for month, totals in sorted(monthly_income_expense.items()):
        ws_summary.append([
            month,
            totals["income"],
            totals["expense_clean"],
            totals["expense_outlier"]
        ])

    ws_summary.auto_filter.ref = ws_summary.dimensions
    add_borders(ws_summary)
    auto_resize_columns(ws_summary)

    # Conditional formatting only up to current month
    current_month_str = datetime.now().strftime("%Y%m")
    cutoff_row = 1
    for row in range(2, ws_summary.max_row + 1):
        cell_value = ws_summary.cell(row=row, column=1).value
        if cell_value and cell_value <= current_month_str:
            cutoff_row = row

    income_range = f"B2:B{cutoff_row}"
    expense_clean_range = f"C2:C{cutoff_row}"
    expense_outlier_range = f"D2:D{cutoff_row}"

    green_yellow_red = ColorScaleRule(
        start_type='min', start_color='63BE7B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='F8696B'
    )
    red_yellow_green = ColorScaleRule(
        start_type='min', start_color='F8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'
    )

    ws_summary.conditional_formatting.add(income_range, red_yellow_green)
    ws_summary.conditional_formatting.add(expense_clean_range, green_yellow_red)
    ws_summary.conditional_formatting.add(expense_outlier_range, green_yellow_red)

    # Monthly detail sheets
    for month, rows in sorted(month_data.items()):
        ws = wb.create_sheet(title=month)
        ws.append(headers)

        for row in rows:
            ws.append(row)

        ws.auto_filter.ref = ws.dimensions
        auto_resize_columns(ws)
        add_borders(ws)

        # Expense breakdown by category (non-outliers only)
        expense_by_category = {}
        for row in rows:
            if row[1] == 'expense' and (row[7] is None or row[7] == 0):
                category = row[2]
                expense_by_category[category] = expense_by_category.get(category, 0) + row[4]

        # Write expense summary table
        start_col = 10  # Column J (because 9 is now 'quantity')
        start_row = 2
        ws.cell(row=start_row, column=start_col, value="Expense Category").font = header_font
        ws.cell(row=start_row, column=start_col + 1, value="Amount").font = header_font

        for i, (cat, amt) in enumerate(expense_by_category.items(), start=start_row + 1):
            ws.cell(row=i, column=start_col, value=cat)
            ws.cell(row=i, column=start_col + 1, value=amt)

        # Add Pie Chart
        if expense_by_category:
            pie = PieChart()
            pie.title = "Expense Breakdown"
            data_ref = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(expense_by_category))
            label_ref = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(expense_by_category))
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(label_ref)
            ws.add_chart(pie, "L2")

    # Line chart for Summary
    chart = LineChart()
    chart.title = "Income and Expense Over Time"
    chart.style = 13
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Month"

    last_row = ws_summary.max_row
    data = Reference(ws_summary, min_col=2, max_col=3, min_row=1, max_row=last_row)
    categories = Reference(ws_summary, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws_summary.add_chart(chart, "F2")

    # Save file
    wb.save(report_file)

def send_report_to_telegram(report_file):
    url = f"https://api.telegram.org/bot{API_TOKEN}/sendDocument"

    with open(report_file, "rb") as f:
        files = {
            "document": f
        }
        data = {
            "chat_id": ALLOWED_USER_ID
        }

        response = requests.post(url, data=data, files=files)
        response.raise_for_status()  # raises error if request failed

# Main function to execute the script
def main():
    report_file = "transactions_report.xlsx"
    
    # Fetch data from database
    data = fetch_data_from_db(DB_PATH)
    
    # Generate Excel report
    generate_excel_report(data, report_file)

    # Send report to Telegram user
    send_report_to_telegram(report_file)

    # Delete the report file after sending
    os.remove(report_file)

if __name__ == "__main__":
    main()
